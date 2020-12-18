import xlrd
import requests

import json
import openpyxl

# 读取测试用例
filename = 'api_testcases.xlsx'
book = xlrd.open_workbook(filename=filename)
names = book.sheet_names()
# print(names)

table1 = book.sheet_by_index(0)
nrows = table1.nrows
ncols = table1.ncols
# print(nrows,ncols)

keys = table1.row_values(0)
# print(keys)
li = []
for i in range(1,nrows):
    values = table1.row_values(i)
    # print(values)
    api = dict(zip(keys,values))
    li.append(api)

# print(li)

api = li[0]
# print(api)



# 执行测试用例
method = api['method']
url = api['url']
params = api['params']
data = eval(api['data'])
expected_result = api['expected_result']
print(method,url,params,data,expected_result)

r1 = requests.request(method=method,url=url,params=params,data=data,verify=False)
result = r1.json()


if result==eval(expected_result):
    print('%s:pass' %api['case_id'])
else:
    print('%s:fail' % api['case_id'])


# 写入测试结果
import openpyxl,json
from openpyxl.styles import Alignment,Font

align = Alignment(wrap_text=True)

wb = openpyxl.load_workbook(filename)
ws = wb.active

ws.cell(2,8,json.dumps(result))
ws['H'].alignment = align
wb.save(filename)


